"""
회사명 하나로 항목1~4,6,8(리스크 일부)을 한 번에 처리하는 파이프라인.

사용법:
    python run_pipeline.py 회사명
    python run_pipeline.py 회사명 --corp-code 01234567   # 동명이인 여러 건일 때 특정
    python run_pipeline.py 회사명 --skip-nts             # 국세청 API가 불안정할 때 건너뛰기

내부적으로 하는 일:
    1. CORPCODE.xml에서 corp_code 검색 (항목1)
    2. 기업개황 조회 (항목1)
    3. 공시검색(F001)으로 감사보고서 존재 확인 (항목2)
    4. 감사보고서 원문(document.xml) 다운로드 + 재무제표/주석 파싱 (항목3,4,6)
    5. 전금업_등록말소현황.xlsx가 같은 폴더에 있으면 라이선스 대조 (항목1,5)
    6. 공시검색(F005)으로 미제출신고 확인 (항목8)
    7. 국세청 사업자상태조회 (항목1,8) - 서버가 불안정해서 실패해도 계속 진행
    8. profile_{corp_code}.json 으로 저장

전제: corp_lookup.py, parse_financials.py, parse_notes.py, nts_lookup.py,
      CORPCODE.xml 이 이 스크립트와 같은 폴더에 있어야 함.
"""
import argparse
import io
import json
import re
import sys
import urllib.request
import zipfile
from pathlib import Path

from corp_lookup import (
    load_corp_index,
    search_by_name,
    fetch_company_overview,
    fetch_disclosure_list,
    check_audit_report_availability,
)
from parse_financials import parse_financial_statements
from parse_notes import load_root, extract_note_section
from nts_lookup import check_business_status, SERVICE_KEY as NTS_KEY

DART_KEY = "053284b30c287dbfd3324294b588649e2729bdca"
BASE_DIR = Path(__file__).parent
LICENSE_XLSX = BASE_DIR / "전금업_등록말소현황.xlsx"

_PREFIX = re.compile(r"^[IVXⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ0-9\.\-]+")
_FIN_WANT = {
    "revenue": ("손 익 계 산 서", ["매출액", "영업수익"]),
    "operating_income": ("손 익 계 산 서", ["영업이익", "영업손실", "영업이익(손실)"]),
    "net_income": ("손 익 계 산 서", ["당기순이익", "당기순손실", "당기순이익(손실)"]),
    "operating_cash_flow": ("현 금 흐 름 표", ["영업활동으로인한현금흐름"]),
    "total_assets": ("재 무 상 태 표", ["자산총계"]),
    "total_liabilities": ("재 무 상 태 표", ["부채총계"]),
    "total_equity": ("재 무 상 태 표", ["자본총계"]),
}


def _norm(s: str) -> str:
    return re.sub(r"\s+", "", s)


def _parse_amount(raw, label=""):
    if raw is None or raw in ("", "-"):
        return None
    neg_paren = raw.startswith("(") and raw.endswith(")")
    num = raw.strip("()").replace(",", "")
    try:
        n = int(num)
    except ValueError:
        return None
    n = -n if neg_paren else n
    if "손실" in label and "이익" not in label and n > 0:
        n = -n
    return n


def extract_financials(xml_text: str) -> dict:
    stm = parse_financial_statements(xml_text)
    out = {}
    for key, (title, patterns) in _FIN_WANT.items():
        rows = stm.get(title, {"rows": []})["rows"]
        found_label, found_val = None, None
        for r in rows:
            core = _PREFIX.sub("", _norm(r["label"]))
            if any(core == _norm(p) or core.startswith(_norm(p)) for p in patterns):
                found_label, found_val = r["label"], (r["values"][0] if r["values"] else None)
                break
        out[key] = _parse_amount(found_val, found_label or "")
    out["unit"] = "KRW"
    out["fiscal_periods"] = stm.get("재 무 상 태 표", {}).get("periods", [])
    return out


def fetch_document_xml(rcept_no: str, dest_dir: Path) -> str:
    dest_dir.mkdir(exist_ok=True)
    url = f"https://opendart.fss.or.kr/api/document.xml?crtfc_key={DART_KEY}&rcept_no={rcept_no}"
    with urllib.request.urlopen(url) as resp:
        data = resp.read()
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        name = zf.namelist()[0]
        zf.extractall(dest_dir)
        return (dest_dir / name).read_text(encoding="utf-8")


def lookup_license(bizr_no: str) -> dict:
    if not LICENSE_XLSX.exists():
        return {"registered": None, "note": "전금업_등록말소현황.xlsx 파일 없음"}
    import openpyxl

    wb = openpyxl.load_workbook(LICENSE_XLSX, data_only=True)
    ws = wb["전자금융업 등록 현황"]
    header = [c.value for c in ws[5]]
    license_cols = header[5:10]
    target = re.sub(r"\D", "", bizr_no)
    for row in ws.iter_rows(min_row=6, values_only=True):
        raw = row[4]
        if raw and re.sub(r"\D", "", str(raw)) == target:
            licenses = [h.replace("\n", " ") for h, v in zip(license_cols, row[5:10]) if v]
            return {
                "registered_name": row[3].strip() if row[3] else None,
                "registration_date": row[2].strftime("%Y-%m-%d") if hasattr(row[2], "strftime") else str(row[2]),
                "licenses": licenses,
            }
    return {"registered": False}


def run(company_name: str, corp_code_override: str | None = None, skip_nts: bool = False) -> dict:
    if corp_code_override:
        corp_code = corp_code_override
    else:
        corps = load_corp_index()
        matches = search_by_name(corps, company_name)
        if len(matches) == 0:
            print(f"'{company_name}' 검색 결과 없음")
            sys.exit(1)
        if len(matches) > 1:
            print(f"'{company_name}' 매칭 {len(matches)}건 - --corp-code로 특정해주세요:")
            for c in matches:
                print(f"  {c['corp_code']}  {c['corp_name']}")
            sys.exit(1)
        corp_code = matches[0]["corp_code"]

    print(f"[1/6] 기업개황 조회 (corp_code={corp_code})")
    overview = fetch_company_overview(corp_code, DART_KEY)

    print("[2/6] 감사보고서(F001) 확인")
    audit = check_audit_report_availability(corp_code, DART_KEY)
    reports = audit.pop("audit_reports")
    if not audit["has_audit_report"]:
        print("  감사보고서 없음 - 재무정보/주석 단계 건너뜀")
        financials, notes, rcept_no = {}, {}, None
    else:
        rcept_no = reports[0]["rcept_no"]
        print(f"[3/6] 감사보고서 원문 다운로드 (rcept_no={rcept_no})")
        xml_text = fetch_document_xml(rcept_no, BASE_DIR / f"{company_name}_doc")
        print("[4/6] 재무제표 + 주석 파싱")
        financials = extract_financials(xml_text)
        root = load_root(str(BASE_DIR / f"{company_name}_doc" / f"{rcept_no}_00760.xml"))
        notes = {
            "company_overview_and_shareholders": extract_note_section(root, "개요", max_chars=1500)
            or extract_note_section(root, "회사", max_chars=1500),
            "related_party_transactions": extract_note_section(root, "특수관계자", max_chars=1500),
        }

    print("[5/6] F005 미제출신고 + 전금업 라이선스 확인")
    f005 = fetch_disclosure_list(corp_code, DART_KEY, pblntf_detail_ty="F005", bgn_de="20180101")
    license_info = lookup_license(overview.get("bizr_no", ""))

    nts_status = None
    if not skip_nts:
        print("[6/6] 국세청 사업자상태 조회 (실패해도 계속 진행)")
        try:
            nts = check_business_status([overview["bizr_no"]], NTS_KEY)
            if nts.get("status_code") == "OK" and nts.get("data"):
                d = nts["data"][0]
                nts_status = {"b_stt": d["b_stt"], "b_stt_cd": d["b_stt_cd"], "tax_type": d["tax_type"], "end_dt": d["end_dt"]}
        except Exception as e:
            print(f"  국세청 API 실패(무시하고 진행): {e}")

    profile = {
        "corp_code": corp_code,
        "basic_info": {
            "corp_name": overview.get("corp_name"),
            "ceo_nm": overview.get("ceo_nm"),
            "adres": overview.get("adres"),
            "bizr_no": overview.get("bizr_no"),
            "jurir_no": overview.get("jurir_no"),
            "corp_cls": overview.get("corp_cls"),
            "induty_code": overview.get("induty_code"),
            "est_dt": overview.get("est_dt"),
            "acc_mt": overview.get("acc_mt"),
            "hm_url": overview.get("hm_url"),
        },
        "electronic_finance_license": license_info,
        "audit_report": {
            "has_audit_report": audit["has_audit_report"],
            "audit_report_count": audit["audit_report_count"],
            "latest_rcept_no": rcept_no,
        },
        "risk_signals": {
            "f005_non_submission_count": len(f005.get("list", [])),
            "nts_business_status": nts_status,
        },
        "financials": financials,
        "notes": notes,
    }

    out_path = BASE_DIR / f"profile_{corp_code}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump({company_name: profile}, f, ensure_ascii=False, indent=2)
    print(f"\n저장 완료: {out_path}")
    return profile


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("company", help="검색할 회사명")
    parser.add_argument("--corp-code", help="동명이인일 때 corp_code 직접 지정")
    parser.add_argument("--skip-nts", action="store_true", help="국세청 API 건너뛰기")
    args = parser.parse_args()

    result = run(args.company, args.corp_code, args.skip_nts)
    print(json.dumps(result, ensure_ascii=False, indent=2))
